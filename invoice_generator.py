"""
Invoice Generator - Automatically generates PDF invoices from Excel data
"""

import os
from datetime import datetime
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import json


class InvoiceGenerator:
    """Generate professional PDF invoices from Excel data"""
    
    def __init__(self, config_file='config.json'):
        """Initialize the invoice generator with configuration"""
        self.load_config(config_file)
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
        
    def load_config(self, config_file):
        """Load configuration from JSON file"""
        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                config = json.load(f)
        else:
            # Default configuration
            config = {
                "company_name": "Your Company Name",
                "company_address": "123 Business Street\nCity, State 12345",
                "company_phone": "+1 (555) 123-4567",
                "company_email": "info@yourcompany.com",
                "logo_path": "",
                "output_folder": "generated_invoices",
                "currency_symbol": "$",
                "thank_you_note": "Thank you for your business!"
            }
            
        self.company_name = config.get('company_name', 'Your Company Name')
        self.company_address = config.get('company_address', '')
        self.company_phone = config.get('company_phone', '')
        self.company_email = config.get('company_email', '')
        self.logo_path = config.get('logo_path', '')
        self.output_folder = config.get('output_folder', 'generated_invoices')
        self.currency_symbol = config.get('currency_symbol', '$')
        self.thank_you_note = config.get('thank_you_note', 'Thank you for your business!')
        
        # Create output folder if it doesn't exist
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
    
    def setup_custom_styles(self):
        """Set up custom paragraph styles"""
        self.styles.add(ParagraphStyle(
            name='CompanyName',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=6,
            alignment=TA_CENTER
        ))
        
        self.styles.add(ParagraphStyle(
            name='CompanyInfo',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#7F8C8D'),
            alignment=TA_CENTER,
            spaceAfter=20
        ))
        
        self.styles.add(ParagraphStyle(
            name='InvoiceTitle',
            parent=self.styles['Heading1'],
            fontSize=28,
            textColor=colors.HexColor('#E74C3C'),
            spaceAfter=12,
            alignment=TA_RIGHT
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#34495E'),
            spaceAfter=6
        ))
    
    def read_excel_data(self, excel_file):
        """Read invoice data from Excel file"""
        wb = load_workbook(excel_file)
        sheet = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Read all invoice data
        invoices = []
        current_invoice = None
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            
            # Check if this is a new invoice (has invoice number)
            if row_data.get('Invoice Number'):
                # Save previous invoice if exists
                if current_invoice:
                    invoices.append(current_invoice)
                
                # Start new invoice
                current_invoice = {
                    'customer_name': row_data.get('Customer Name', ''),
                    'address': row_data.get('Address', ''),
                    'phone': row_data.get('Phone Number', ''),
                    'invoice_number': str(row_data.get('Invoice Number', '')),
                    'date': row_data.get('Date', datetime.now().strftime('%Y-%m-%d')),
                    'tax_percent': float(row_data.get('Tax %', 0) or 0),
                    'discount_percent': float(row_data.get('Discount %', 0) or 0),
                    'items': []
                }
            
            # Add item to current invoice
            if current_invoice and row_data.get('Item Name'):
                item = {
                    'name': row_data.get('Item Name', ''),
                    'quantity': float(row_data.get('Quantity', 0) or 0),
                    'price': float(row_data.get('Price', 0) or 0)
                }
                current_invoice['items'].append(item)
        
        # Add last invoice
        if current_invoice:
            invoices.append(current_invoice)
        
        wb.close()
        return invoices
    
    def calculate_totals(self, items, tax_percent, discount_percent):
        """Calculate invoice totals"""
        subtotal = sum(item['quantity'] * item['price'] for item in items)
        discount_amount = subtotal * (discount_percent / 100)
        subtotal_after_discount = subtotal - discount_amount
        tax_amount = subtotal_after_discount * (tax_percent / 100)
        total = subtotal_after_discount + tax_amount
        
        return {
            'subtotal': subtotal,
            'discount_amount': discount_amount,
            'subtotal_after_discount': subtotal_after_discount,
            'tax_amount': tax_amount,
            'total': total
        }
    
    def generate_pdf(self, invoice_data):
        """Generate a single PDF invoice"""
        invoice_number = invoice_data['invoice_number']
        filename = os.path.join(self.output_folder, f"Invoice_{invoice_number}.pdf")
        
        doc = SimpleDocTemplate(
            filename,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        story = []
        
        # Add logo if available
        if self.logo_path and os.path.exists(self.logo_path):
            logo = Image(self.logo_path, width=1.5*inch, height=1.5*inch)
            story.append(logo)
            story.append(Spacer(1, 0.2*inch))
        
        # Company header
        story.append(Paragraph(self.company_name, self.styles['CompanyName']))
        
        company_info = f"{self.company_address}<br/>"
        if self.company_phone:
            company_info += f"Phone: {self.company_phone}<br/>"
        if self.company_email:
            company_info += f"Email: {self.company_email}"
        
        story.append(Paragraph(company_info, self.styles['CompanyInfo']))
        story.append(Spacer(1, 0.3*inch))
        
        # Invoice title and number
        story.append(Paragraph("INVOICE", self.styles['InvoiceTitle']))
        
        # Invoice details table
        invoice_date = invoice_data['date']
        if isinstance(invoice_date, datetime):
            invoice_date = invoice_date.strftime('%Y-%m-%d')
        elif invoice_date:
            invoice_date = str(invoice_date)
        else:
            invoice_date = datetime.now().strftime('%Y-%m-%d')
        
        details_data = [
            ['Invoice Number:', invoice_number, 'Date:', invoice_date],
        ]
        
        details_table = Table(details_data, colWidths=[1.5*inch, 2*inch, 0.8*inch, 1.5*inch])
        details_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2C3E50')),
        ]))
        
        story.append(details_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Customer information
        story.append(Paragraph("Bill To:", self.styles['SectionHeader']))
        
        customer_info = f"<b>{invoice_data['customer_name']}</b><br/>"
        if invoice_data['address']:
            customer_info += f"{invoice_data['address']}<br/>"
        if invoice_data['phone']:
            customer_info += f"Phone: {invoice_data['phone']}"
        
        story.append(Paragraph(customer_info, self.styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Items table
        items_data = [['Item', 'Quantity', 'Price', 'Total']]
        
        for item in invoice_data['items']:
            line_total = item['quantity'] * item['price']
            items_data.append([
                item['name'],
                f"{item['quantity']:.2f}",
                f"{self.currency_symbol}{item['price']:.2f}",
                f"{self.currency_symbol}{line_total:.2f}"
            ])
        
        items_table = Table(items_data, colWidths=[3.5*inch, 1*inch, 1*inch, 1.3*inch])
        items_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(items_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Calculate totals
        totals = self.calculate_totals(
            invoice_data['items'],
            invoice_data['tax_percent'],
            invoice_data['discount_percent']
        )
        
        # Totals table
        totals_data = []
        
        # Subtotal
        totals_data.append(['Subtotal:', f"{self.currency_symbol}{totals['subtotal']:.2f}"])
        
        # Discount if applicable
        if invoice_data['discount_percent'] > 0:
            totals_data.append([
                f"Discount ({invoice_data['discount_percent']:.1f}%):",
                f"-{self.currency_symbol}{totals['discount_amount']:.2f}"
            ])
            totals_data.append([
                'Subtotal after Discount:',
                f"{self.currency_symbol}{totals['subtotal_after_discount']:.2f}"
            ])
        
        # Tax if applicable
        if invoice_data['tax_percent'] > 0:
            totals_data.append([
                f"Tax ({invoice_data['tax_percent']:.1f}%):",
                f"{self.currency_symbol}{totals['tax_amount']:.2f}"
            ])
        
        # Total
        totals_data.append(['Total Amount Due:', f"{self.currency_symbol}{totals['total']:.2f}"])
        
        totals_table = Table(totals_data, colWidths=[4.8*inch, 1.9*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -2), 10),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#E74C3C')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#34495E')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(totals_table)
        story.append(Spacer(1, 0.5*inch))
        
        # Thank you note
        thank_you_style = ParagraphStyle(
            'ThankYou',
            parent=self.styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#27AE60'),
            alignment=TA_CENTER,
            spaceAfter=10
        )
        story.append(Paragraph(self.thank_you_note, thank_you_style))
        
        # Footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#95A5A6'),
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            f"This invoice was generated automatically on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            footer_style
        ))
        
        # Build PDF
        doc.build(story)
        return filename
    
    def process_excel_file(self, excel_file):
        """Process Excel file and generate all invoices"""
        print(f"Reading invoice data from: {excel_file}")
        invoices = self.read_excel_data(excel_file)
        
        print(f"\nFound {len(invoices)} invoice(s) to generate")
        
        generated_files = []
        for idx, invoice_data in enumerate(invoices, 1):
            print(f"\nGenerating invoice {idx}/{len(invoices)}: {invoice_data['invoice_number']}")
            filename = self.generate_pdf(invoice_data)
            generated_files.append(filename)
            print(f"  ✓ Created: {filename}")
        
        print(f"\n{'='*60}")
        print(f"Successfully generated {len(generated_files)} invoice(s)")
        print(f"Output folder: {os.path.abspath(self.output_folder)}")
        print(f"{'='*60}")
        
        return generated_files


def main():
    """Main function to run the invoice generator"""
    import sys
    
    print("="*60)
    print("Invoice Generator - Excel to PDF")
    print("="*60)
    
    # Check command line arguments
    if len(sys.argv) < 2:
        print("\nUsage: python invoice_generator.py <excel_file>")
        print("Example: python invoice_generator.py sample_invoices.xlsx")
        print("\nOr edit this script to specify the Excel file directly.")
        
        # Default file for testing
        excel_file = "sample_invoices.xlsx"
        if not os.path.exists(excel_file):
            print(f"\nError: Sample file '{excel_file}' not found.")
            print("Please provide an Excel file as argument or create sample_invoices.xlsx")
            return
    else:
        excel_file = sys.argv[1]
    
    # Check if file exists
    if not os.path.exists(excel_file):
        print(f"\nError: File '{excel_file}' not found!")
        return
    
    # Generate invoices
    generator = InvoiceGenerator()
    generator.process_excel_file(excel_file)


if __name__ == "__main__":
    main()
