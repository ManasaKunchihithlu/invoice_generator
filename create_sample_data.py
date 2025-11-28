"""
Sample Invoice Data Generator
Creates a sample Excel file with invoice data for testing
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random


def create_sample_excel():
    """Create a sample Excel file with invoice data"""
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Invoices"
    
    # Define headers
    headers = [
        'Invoice Number',
        'Customer Name',
        'Address',
        'Phone Number',
        'Date',
        'Item Name',
        'Quantity',
        'Price',
        'Tax %',
        'Discount %'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sample data arrays for generating random invoices
    customer_names = [
        'John Smith', 'Sarah Johnson', 'Tech Solutions Inc', 'Mike Davis', 'Emily Brown',
        'Global Systems Ltd', 'David Wilson', 'Jennifer Martinez', 'ABC Corporation',
        'Robert Taylor', 'Lisa Anderson', 'Innovation Labs', 'James Thomas', 'Mary Jackson',
        'Digital Partners LLC', 'Michael White', 'Patricia Harris', 'Enterprise Group',
        'Christopher Martin', 'Linda Thompson', 'Smart Tech Inc', 'Daniel Garcia',
        'Elizabeth Rodriguez', 'Cloud Services Co', 'Matthew Robinson'
    ]
    
    addresses = [
        '123 Main St, New York, NY 10001', '456 Oak Avenue, Los Angeles, CA 90001',
        '789 Business Blvd, Suite 100, Chicago, IL 60601', '321 Elm Street, Houston, TX 77001',
        '654 Pine Road, Phoenix, AZ 85001', '987 Maple Drive, Philadelphia, PA 19019',
        '147 Cedar Lane, San Antonio, TX 78201', '258 Birch Way, San Diego, CA 92101',
        '369 Walnut Court, Dallas, TX 75201', '741 Spruce Avenue, San Jose, CA 95101',
        '852 Ash Boulevard, Austin, TX 78701', '963 Oak Street, Jacksonville, FL 32099',
        '159 Pine Circle, Fort Worth, TX 76101', '357 Maple Terrace, Columbus, OH 43004',
        '486 Cedar Plaza, Charlotte, NC 28201'
    ]
    
    items = [
        'Laptop Computer', 'Wireless Mouse', 'USB-C Cable', 'Office Desk', 'Office Chair',
        'Desk Lamp', 'Web Development Service', 'SEO Optimization', 'Logo Design',
        'Monitor 27"', 'Keyboard Mechanical', 'External Hard Drive', 'Printer', 'Scanner',
        'Webcam HD', 'Headphones', 'Microphone', 'Docking Station', 'Cable Organizer',
        'Standing Desk', 'Ergonomic Mat', 'Whiteboard', 'Projector', 'Conference Phone',
        'Router', 'Network Switch', 'UPS Battery Backup', 'Server Rack', 'Consulting Services',
        'Software License', 'Cloud Storage', 'Technical Support', 'Training Session',
        'Marketing Campaign', 'Graphic Design', 'Content Writing', 'Video Editing',
        'Data Analysis', 'Security Audit', 'Mobile App Development'
    ]
    
    # Generate 200 invoices (this will create 200-600 rows depending on items per invoice)
    row_num = 2
    invoice_count = 0
    target_invoices = 200
    
    while invoice_count < target_invoices:
        invoice_count += 1
        invoice_num = f'INV-{invoice_count:03d}'
        customer = random.choice(customer_names)
        address = random.choice(addresses)
        phone = f'+1-555-{random.randint(1000, 9999)}'
        invoice_date = (datetime.now() - timedelta(days=random.randint(0, 90))).strftime('%Y-%m-%d')
        
        # Each invoice has 1-3 items
        num_items = random.randint(1, 3)
        
        for item_idx in range(num_items):
            if invoice_count > target_invoices:  # Safety check
                break
                
            item_name = random.choice(items)
            quantity = random.randint(1, 50)
            price = round(random.uniform(50, 5000), 2)
            
            # First item of invoice has full details, subsequent items are blank for invoice info
            if item_idx == 0:
                tax = round(random.uniform(5, 10), 1)
                discount = random.choice([0, 5, 10, 15, 20])
                row_data = [invoice_num, customer, address, phone, invoice_date, 
                           item_name, quantity, price, tax, discount]
            else:
                row_data = ['', '', '', '', '', item_name, quantity, price, '', '']
            
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
            
            row_num += 1
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # Invoice Number
        'B': 25,  # Customer Name
        'C': 40,  # Address
        'D': 18,  # Phone Number
        'E': 12,  # Date
        'F': 30,  # Item Name
        'G': 12,  # Quantity
        'H': 12,  # Price
        'I': 10,  # Tax %
        'J': 12,  # Discount %
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # Save the workbook
    filename = 'sample_invoices.xlsx'
    wb.save(filename)
    print(f"Sample Excel file created: {filename}")
    print(f"\nThis file contains {row_num - 2} rows of sample invoice data with {invoice_count} invoices.")
    print(f"Each invoice has 1-3 line items.")
    print("You can use this as a template for your own invoice data.")
    
    return filename


if __name__ == "__main__":
    create_sample_excel()
